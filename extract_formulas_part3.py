
import openpyxl
import os

file_path = r'src/Cotizador impresión 3D  V 1.0 2023.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

sheet_name = 'Desglose de calculo '
if sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} (Rows 20-30) ---")
    sheet = wb[sheet_name]
    
    for row in range(20, 31):
        row_data = []
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                row_data.append(f"{cell.coordinate}: {value}")
        
        if row_data:
            print(" | ".join(row_data))
else:
    print(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
