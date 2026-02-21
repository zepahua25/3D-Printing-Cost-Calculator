
import openpyxl
import os

file_path = r'src/Cotizador impresión 3D  V 1.0 2023.xlsx'

print(f"Reading file: {file_path}")

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except FileNotFoundError:
    print(f"Error: File not found at {file_path}")
    print(f"Current working directory: {os.getcwd()}")
    exit(1)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    sheet = wb[sheet_name]
    
    # Print the first 20 rows and 10 columns
    for row in range(1, 21):
        row_data = []
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                row_data.append(f"{cell.coordinate}: {value}")
        
        if row_data:
            print(" | ".join(row_data))
